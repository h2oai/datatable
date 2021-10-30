#!/usr/bin/env python3
#-------------------------------------------------------------------------------
# Copyright 2018-2021 H2O.ai
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#-------------------------------------------------------------------------------
import datatable as dt



def read_xlsx_workbook(filename, subpath):
    try:
        import openpyxl
    except ImportError:
        raise dt.exceptions.ImportError(
            "Module `openpyxl` is required in order to read Excel file '%s'"
            % filename)

    if subpath:
        wb = openpyxl.load_workbook(filename, data_only=True)
        range2d = None
        if subpath in wb.get_sheet_names():
            sheetname = subpath
        else:
            if "/" in subpath:
                sheetname, xlsrange = subpath.rsplit('/', 1)
                range2d = xlsrange
            if not(sheetname in wb.sheetnames and range2d is not None):
                raise ValueError("Sheet `%s` is not found in the XLS file"
                                 % subpath)
        ws = wb.get_sheet_by_name(sheetname)
        result = read_xlsx_worksheet(ws, range2d)
    else:
        wb = openpyxl.load_workbook(filename, data_only=True)
        result = {}
        for name, ws in zip(wb.sheetnames, wb):
            out = read_xlsx_worksheet(ws)
            if out is None:
                continue
            for i, frame in out.items():
                result["%s/%s/%s" % (filename, name, i)] = frame

    if len(result) == 0:
        return None
    elif len(result) == 1:
        for v in result.values():
            return v
    else:
        return result



def read_xlsx_worksheet(ws, subrange=None):
    if subrange is None:
        ranges2d = [ws.calculate_dimension()]
    else:
        ranges2d = [subrange]

    results = {}
    for range2d in ranges2d:
        subview = ws[range2d]
        # the subview is a tuple of rows, which are tuples of columns
        ncols = len(subview[0])
        nrows = len(subview) - 1
        colnames = [str(n.value) for n in subview[0]]
        coltypes = [None] * ncols
        outdata = [[None] * nrows for _ in range(ncols)]
        for irow, row in enumerate(subview[1:]):
            for icol, cell in enumerate(row):
                cell_value = cell.value
                cell_type = cell.data_type
                if not cell_type or cell_type == "e" or cell_value is None:
                    cell_value = None
                    cell_type = None
                elif cell_type == "b":
                    cell_value = bool(cell_value)
                    cell_type = bool
                else:
                    # float, int, date, etc
                    cell_type = type(cell.value)

                outdata[icol][irow] = cell_value
                prev_type = coltypes[icol]
                if prev_type is None and cell_type is not None:
                    coltypes[icol] = cell_type
                elif cell_type is None or cell_type == prev_type:
                    pass
                elif cell_type is str:
                    # str always win
                    coltypes[icol] = str

        # let the frame decide on the types of non-str cols
        types = [str if coltype == str else None
                 for coltype in coltypes]
        frame = dt.Frame(outdata, names=colnames, types=types)
        results[range2d] = frame
    return results
